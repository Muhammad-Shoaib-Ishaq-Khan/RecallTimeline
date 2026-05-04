"""
╔══════════════════════════════════════════════════════════════════╗
║   RecallTimeline Windows Recall Forensic Analysis Tool          ║
║   Authors : Muhammad Shoaib Ishaq Khan, Ahmad Hassan,           ║
║             Zoha Nazar | Digital Forensics Research             ║
║   Supervisor: Dr. Ali Sufiyan                                   ║
║   License : MIT  |  Python : 3.9+                               ║
╚══════════════════════════════════════════════════════════════════╝

SOURCES SUPPORTED:
  1. ukg.db        Real Windows Recall SQLite database
  2. ImageStore/   Real Recall JPEG EXIF MakerNote (tag 0x927C)
  3. Flat folder   capture_snapshots.py JPEG output

USAGE:
    python RecallTimeline.py analyse --recall-dir ./snapshots --case MyCase
    pyinstaller --onefile --name RecallTimeline RecallTimeline.py
"""

import os, sys, csv, json, struct, hashlib, sqlite3
import argparse, datetime, random, string
from pathlib import Path
import html as html_lib

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────
TOOL_NAME    = "RecallTimeline"
TOOL_AUTHOR  = "M.S.I. Khan | A. Hassan | Z. Nazar"
MAKERNOTE_TAG = 0x927C

RECALL_FIELDS = [
    "Timestamp", "WindowTitle", "ProcessPath",
    "URL", "IsForeground", "SnapshotId", "AppId",
]

# Each entry: (keyword_to_match, human_readable_reason)
SUSPICIOUS_RULES = [
    ("confidential", "File/window marked CONFIDENTIAL — possible sensitive data access"),
    ("secret",       "File/window contains 'secret' — possible sensitive data access"),
    ("private",      "File/window marked PRIVATE — restricted data indicator"),
    ("exfil",        "Keyword 'exfil' detected — data exfiltration indicator"),
    ("export",       "Export action detected — data being copied/exported"),
    ("upload",       "Upload action detected — data being sent externally"),
    ("transfer",     "Transfer action detected — data movement indicator"),
    ("pastebin",     "Pastebin.com detected — common code/data leak platform"),
    ("mega.nz",      "MEGA cloud storage detected — known data exfiltration platform"),
    ("wetransfer",   "WeTransfer detected — large file sharing platform"),
    ("dropbox",      "Dropbox detected — cloud file sync/share platform"),
    ("cmd.exe",      "Windows Command Prompt (cmd.exe) — command execution detected"),
    ("powershell",   "PowerShell execution detected — scripting/admin activity"),
    ("wscript",      "Windows Script Host detected — script execution indicator"),
    ("taskmgr",      "Task Manager opened — possible process termination/hiding attempt"),
    ("regedit",      "Registry Editor detected — system configuration modification"),
    ("mimikatz",     "Mimikatz keyword detected — credential harvesting tool"),
    ("procdump",     "ProcDump detected — memory dump/credential extraction tool"),
    ("netcat",       "Netcat detected — network tunneling/reverse shell tool"),
    ("ngrok",        "Ngrok detected — tunneling tool, possible C2 channel"),
    ("base64",       "Base64 encoding detected — common obfuscation technique"),
    ("certutil",     "CertUtil detected — can be abused for file download/decoding"),
    ("bitsadmin",    "BITSAdmin detected — can be abused for stealthy downloads"),
    ("mshta",        "MSHTA detected — HTA execution, common malware vector"),
    ("regsvr32",     "Regsvr32 detected — DLL registration abuse technique"),
    ("python",       "Python interpreter running — script execution in terminal"),
    ("wget",         "Wget detected — command-line file download tool"),
    ("curl",         "Curl detected — command-line data transfer tool"),
    ("ftp",          "FTP activity detected — file transfer protocol usage"),
]

SUSPICIOUS_KEYWORDS = [r[0] for r in SUSPICIOUS_RULES]
FLAG_REASONS        = {r[0]: r[1] for r in SUSPICIOUS_RULES}

BANNER = f"""
				╔══════════════════════════════════════════════════════════════════╗
				║  {TOOL_NAME} Windows Recall Forensic Tool                     ║
				║  Authors: M.S.I. Khan, A. Hassan, Z. Nazar                       ║
				║  Supervisor: Dr. Ali Sufiyan                                     ║
				║  For authorized forensic use only                                ║
				╚══════════════════════════════════════════════════════════════════╝


                                ╔═══════════════════════════════════════════════════════════════════╗                                       
                                ║                                                                   ║                                              
                                ║    ██████╗   ███████╗  ██████╗  █████╗   ██╗     ██╗              ║                                          
                                ║    ██╔══██╗  ██╔════╝ ██╔════╝  ██╔══██╗ ██║     ██║              ║                                           
                                ║    ██████╔╝  █████╗   ██║       ███████║ ██║     ██║              ║                                          
                                ║    ██╔══██╗  ██╔══╝   ██║       ██╔══██║ ██║     ██║              ║                                           
                                ║    ██║  ██║  ███████╗╚██████╗   ██║  ██║ ███████╗███████╗         ║                                             
                                ║    ╚═╝  ╚═╝  ╚══════╝ ╚═════╝   ╚═╝  ╚═╝ ╚══════╝╚══════╝         ║                                         
                                ║                                                                   ║                                        
                                ║ ████████╗██╗███╗   ███╗███████╗██╗     ██╗ ███╗   ██╗███████╗     ║                                          
                                ║ ╚══██╔══╝██║████╗ ████║██╔════╝██║     ██║ ████╗  ██║██╔════╝     ║                                          
                                ║    ██║   ██║██╔████╔██║█████╗  ██║     ██║ ██╔██╗ ██║█████╗       ║                                                
                                ║    ██║   ██║██║╚██╔╝██║██╔══╝  ██║     ██║ ██║╚██╗██║██╔══╝       ║                                                
                                ║    ██║   ██║██║ ╚═╝ ██║███████╗███████╗██║ ██║ ╚████║███████╗     ║                                                
                                ║    ╚═╝   ╚═╝╚═╝     ╚═╝╚══════╝╚══════╝╚═╝ ╚═╝  ╚═══╝╚══════╝     ║                                                  
                                ║                                                               v1.0║                                             
                                ╚═══════════════════════════════════════════════════════════════════╝      

"""

CSV_FIELDS = [
    "Source", "Timestamp", "WindowTitle", "ProcessPath", "URL",
    "IsForeground", "OcrText", "Filename", "SHA256", "FileSize",
    "MTime", "CTime", "Flags", "FlagReasons", "AutopsyCorroboration",
]

# ══════════════════════════════════════════════════════════════════
#  UTILITY
# ══════════════════════════════════════════════════════════════════

def compute_sha256(filepath: str) -> str:
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _empty_record(source: str) -> dict:
    return {
        "Source": source, "EventId": None,
        "Timestamp": "", "SortKey": None,
        "WindowTitle": "", "ProcessPath": "", "URL": "",
        "IsForeground": "", "SnapshotId": "", "AppId": "",
        "OcrText": "", "Filename": "", "FilePath": "",
        "FileSize": 0, "SHA256": "", "MTime": "", "CTime": "",
        "Flags": "", "FlagReasons": "", "AutopsyCorroboration": "",
    }


# ══════════════════════════════════════════════════════════════════
#  PARSERS
# ══════════════════════════════════════════════════════════════════

def parse_ukg_db(db_path: str) -> list[dict]:
    records = []
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cur  = conn.cursor()
        cur.execute("""
            SELECT wc.Id, wc.WindowTitle, wc.TimeStamp, wc.ImageToken,
                   wc.AppId, wc.IsActive, wc.ProcessPath,
                   w.WebUrl, ti.OcrText
            FROM   WindowCapture wc
            LEFT JOIN Web w  ON wc.Id = w.WindowCaptureId
            LEFT JOIN WindowCaptureTextIndex ti ON wc.Id = ti.WindowCaptureId
            WHERE  wc.WindowTitle IS NOT NULL OR wc.ImageToken IS NOT NULL
            ORDER  BY wc.TimeStamp ASC
        """)
        for row in cur.fetchall():
            ts_ms = row["TimeStamp"] or 0
            ts    = datetime.datetime.fromtimestamp(ts_ms / 1000)
            rec   = _empty_record("ukg.db")
            rec.update({
                "EventId":      row["Id"],
                "Timestamp":    ts.isoformat(),
                "SortKey":      ts,
                "WindowTitle":  row["WindowTitle"] or "",
                "ProcessPath":  row["ProcessPath"] or "",
                "URL":          row["WebUrl"] or "",
                "IsForeground": str(bool(row["IsActive"])),
                "SnapshotId":   row["ImageToken"] or "",
                "AppId":        row["AppId"] or "",
                "OcrText":      row["OcrText"] or "",
                "Filename":     (row["ImageToken"] or "") + ".jpg",
            })
            records.append(rec)
        conn.close()
    except sqlite3.Error as e:
        print(f"  [WARN] SQLite error: {e}")
    return records


def _decode_makernote(raw: bytes) -> dict:
    rec = {k: "" for k in RECALL_FIELDS}
    if not raw:
        return rec

    # Strategy 1: binary key-value (production Recall)
    try:
        pos, parsed = 0, {}
        while pos + 8 < len(raw):
            kt = struct.unpack_from("<I", raw, pos)[0]; pos += 4
            if kt != 0x0C:
                break
            kl = struct.unpack_from("<I", raw, pos)[0]; pos += 4
            if kl <= 0 or pos + kl > len(raw):
                break
            key = raw[pos:pos+kl].decode("utf-8", errors="replace"); pos += kl
            vt  = struct.unpack_from("<I", raw, pos)[0]; pos += 4
            if   vt == 0x0C:
                vl  = struct.unpack_from("<I", raw, pos)[0]; pos += 4
                val = raw[pos:pos+vl].decode("utf-8", errors="replace"); pos += vl
            elif vt == 0x04: val = str(struct.unpack_from("<I", raw, pos)[0]); pos += 4
            elif vt == 0x08: val = str(struct.unpack_from("<Q", raw, pos)[0]); pos += 8
            elif vt == 0x01: val = str(bool(struct.unpack_from("<B", raw, pos)[0]));  pos += 1
            else: break
            parsed[key] = val
        if parsed:
            for k in RECALL_FIELDS:
                rec[k] = parsed.get(k, "")
            return rec
    except Exception:
        pass

    # Strategy 2: length-prefixed JSON
    try:
        if len(raw) >= 4:
            jlen = struct.unpack_from("<I", raw, 0)[0]
            if 0 < jlen <= len(raw) - 4:
                p = json.loads(raw[4:4+jlen].decode("utf-8", errors="replace"))
                for k in RECALL_FIELDS:
                    rec[k] = str(p.get(k, ""))
                return rec
    except Exception:
        pass

    # Strategy 3: raw UTF-8 JSON
    try:
        text  = raw.decode("utf-8", errors="replace").strip()
        start = text.find("{")
        if start != -1:
            p = json.loads(text[start:])
            for k in RECALL_FIELDS:
                rec[k] = str(p.get(k, ""))
            return rec
    except Exception:
        pass

    # Strategy 4: UTF-16 LE JSON
    try:
        text = raw.decode("utf-16-le", errors="replace")
        if "{" in text:
            p = json.loads(text[text.find("{"):text.rfind("}")+1])
            for k in RECALL_FIELDS:
                rec[k] = str(p.get(k, ""))
            return rec
    except Exception:
        pass

    return rec


def _read_exif_makernote(jpeg_path: str):
    if not PIL_AVAILABLE:
        return None
    try:
        img  = Image.open(jpeg_path)
        exif = img._getexif()
        if exif:
            return exif.get(MAKERNOTE_TAG)
    except Exception:
        pass
    return None


def _read_sidecar_json(jpeg_path: str) -> dict | None:
    base = os.path.splitext(jpeg_path)[0]
    for ext in (".json", "_meta.json"):
        p = base + ext
        if os.path.isfile(p):
            try:
                with open(p, encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
    return None


def _parse_filename_ts(jpeg_path: str) -> dict:
    """Extract timestamp from recall_YYYYMMDD_HHMMSS_NNN.jpg naming."""
    stem  = Path(jpeg_path).stem
    parts = stem.split("_")
    try:
        if len(parts) >= 3 and parts[0] == "recall":
            dt = datetime.datetime.strptime(parts[1] + parts[2], "%Y%m%d%H%M%S")
            return {"Timestamp": dt.isoformat()}
    except Exception:
        pass
    return {}


def parse_jpeg(jpeg_path: str) -> dict | None:
    try:
        stat   = os.stat(jpeg_path)
        sha256 = compute_sha256(jpeg_path)

        raw_maker  = _read_exif_makernote(jpeg_path)
        exif_meta  = _decode_makernote(raw_maker)
        sidecar    = _read_sidecar_json(jpeg_path) or {}
        fname_meta = _parse_filename_ts(jpeg_path)

        def best(field, fallback=""):
            return (exif_meta.get(field) or
                    str(sidecar.get(field, "")) or
                    fname_meta.get(field, "") or
                    fallback)

        ts_str = best("Timestamp")
        if not ts_str:
            ts     = datetime.datetime.fromtimestamp(stat.st_mtime)
            ts_str = ts.isoformat()
        else:
            try:
                ts = datetime.datetime.fromisoformat(ts_str)
            except ValueError:
                ts     = datetime.datetime.fromtimestamp(stat.st_mtime)
                ts_str = ts.isoformat()

        window_title = best("WindowTitle")
        process_path = best("ProcessPath")
        url          = best("URL")
        has_rich     = bool(window_title or process_path)
        has_exif     = raw_maker is not None

        source = ("ImageStore (EXIF)"    if has_rich and has_exif else
                  "Capture (sidecar)"    if has_rich and sidecar   else
                  "Capture (filename)"   if fname_meta.get("Timestamp") else
                  "ImageStore (fs-only)")

        rec = _empty_record(source)
        rec.update({
            "Timestamp":    ts_str,
            "SortKey":      ts,
            "WindowTitle":  window_title,
            "ProcessPath":  process_path,
            "URL":          url,
            "IsForeground": best("IsForeground"),
            "SnapshotId":   best("SnapshotId", Path(jpeg_path).stem),
            "AppId":        best("AppId"),
            "Filename":     os.path.basename(jpeg_path),
            "FilePath":     jpeg_path,
            "FileSize":     stat.st_size,
            "SHA256":       sha256,
            "MTime":        datetime.datetime.fromtimestamp(stat.st_mtime).isoformat(),
            "CTime":        datetime.datetime.fromtimestamp(stat.st_ctime).isoformat(),
        })
        return rec
    except Exception as e:
        print(f"  [WARN] JPEG parse failed: {jpeg_path}: {e}")
        return None


# ══════════════════════════════════════════════════════════════════
#  TIMELINE BUILDER
# ══════════════════════════════════════════════════════════════════

def build_timeline(recall_dir: str, autopsy_csv: str = None) -> list[dict]:
    records, seen_ids = [], set()

    db_path = os.path.join(recall_dir, "ukg.db")
    if os.path.isfile(db_path):
        print(f"[*] Parsing ukg.db → {db_path}")
        db_recs = parse_ukg_db(db_path)
        for rec in db_recs:
            key = rec.get("SnapshotId") or str(rec.get("EventId"))
            if key and key not in seen_ids:
                seen_ids.add(key)
                records.append(rec)
        print(f"    → {len(db_recs)} records from ukg.db")
    else:
        print("  [INFO] No ukg.db — skipping database source")

    _parse_jpeg_folder(os.path.join(recall_dir, "ImageStore"),
                       records, seen_ids, "ImageStore", recurse=True)
    _parse_jpeg_folder(recall_dir, records, seen_ids,
                       "Direct folder", recurse=False)

    if not records:
        print("[!] No artifacts found.")
        return []

    records.sort(key=lambda r: r["SortKey"])

    if autopsy_csv and os.path.isfile(autopsy_csv):
        records = _merge_autopsy(records, autopsy_csv)

    print(f"[*] Timeline built: {len(records)} events")
    return records


def _parse_jpeg_folder(folder, records, seen_ids, label, recurse=True):
    if not os.path.isdir(folder):
        return
    jpegs = (
        list(Path(folder).glob("**/*.jpg" if recurse else "*.jpg")) +
        list(Path(folder).glob("**/*.jpeg" if recurse else "*.jpeg"))
    )
    if recurse:
        jpegs += [p for p in Path(folder).rglob("*")
                  if p.is_file() and not p.suffix]
    print(f"[*] {label}: {len(jpegs)} file(s) → {folder}")
    added = 0
    for path in jpegs:
        rec = parse_jpeg(str(path))
        if rec is None:
            continue
        sid = rec.get("SnapshotId", "")
        if sid and sid in seen_ids:
            continue
        if sid:
            seen_ids.add(sid)
        records.append(rec)
        added += 1
    print(f"    → {added} records added")


def _merge_autopsy(records, autopsy_csv):
    if not autopsy_csv or not os.path.isfile(autopsy_csv):
        return records
    
    autopsy = {}
    ext = os.path.splitext(autopsy_csv)[1].lower()
    
    # Handle .xlsx files
    if ext == '.xlsx':
        if not XLSX_AVAILABLE:
            print("  [WARN] openpyxl not installed. Install with: pip install openpyxl")
            print("  [WARN] Or export as CSV from Autopsy instead of Excel")
            return records
        try:
            wb = openpyxl.load_workbook(autopsy_csv, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                name = str(row_dict.get("Name", "")).lower()
                if name:
                    autopsy[name] = row_dict
            wb.close()
            print(f"    [i] Autopsy XLSX parsed: {len(autopsy)} entries")
        except Exception as e:
            print(f"  [WARN] Autopsy XLSX error: {e}")
            return records
    
    # Handle .csv files
    else:
        for encoding in ["utf-8-sig", "cp1252", "latin-1", "utf-8"]:
            try:
                with open(autopsy_csv, newline="", encoding=encoding) as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        name = str(row.get("Name", "")).lower()
                        if name:
                            autopsy[name] = row
                print(f"    [i] Autopsy CSV parsed ({encoding}): {len(autopsy)} entries")
                break
            except (UnicodeDecodeError, Exception):
                continue
    
    if not autopsy:
        print("  [WARN] Could not parse Autopsy file with any method")
        return records
    
    # Match by filename (exact match first, then partial)
    matched = 0
    for rec in records:
        rec_fname = str(rec.get("Filename", "")).lower()
        rec_proc  = os.path.basename(str(rec.get("ProcessPath", ""))).lower()
        rec_title = str(rec.get("WindowTitle", "")).lower()
        
        rec["AutopsyCorroboration"] = ""
        found = False
        
        # Try exact filename match
        if rec_fname and rec_fname in autopsy:
            e = autopsy[rec_fname]
            rec["AutopsyCorroboration"] = (
                f"FS match: {e.get('Name','')} modified {e.get('DateModified','')}"
            )
            matched += 1
            found = True
        
        # Try process name match
        if not found and rec_proc:
            for aut_name, aut_data in autopsy.items():
                if rec_proc in aut_name:
                    rec["AutopsyCorroboration"] = (
                        f"FS match (process): {aut_data.get('Name','')}"
                    )
                    matched += 1
                    found = True
                    break
        
        # Try window title keywords
        if not found:
            for word in rec_title.split():
                word = word.strip().lower()
                if len(word) > 3 and word in autopsy:
                    rec["AutopsyCorroboration"] = (
                        f"FS match (keyword): {word}"
                    )
                    matched += 1
                    found = True
                    break
    
    print(f"    [i] Autopsy corroboration: {matched}/{len(records)} events matched")
    return records

# ══════════════════════════════════════════════════════════════════
#  ANOMALY DETECTION
# ══════════════════════════════════════════════════════════════════

def flag_anomalies(records: list[dict]) -> list[dict]:
    for rec in records:
        flags, reasons = [], []
        combined = " ".join([
            rec.get("WindowTitle", ""),
            rec.get("URL", ""),
            rec.get("ProcessPath", ""),
            rec.get("OcrText", ""),
        ]).lower()

        for kw in SUSPICIOUS_KEYWORDS:
            if kw in combined:
                flags.append(kw)
                reasons.append(FLAG_REASONS.get(kw, kw))

        rec["Flags"]       = ", ".join(flags)
        rec["FlagReasons"] = " | ".join(reasons)

    # Burst detection
    for i in range(1, len(records)):
        try:
            t1 = records[i-1]["SortKey"]
            t2 = records[i]["SortKey"]
            if isinstance(t1, datetime.datetime) and isinstance(t2, datetime.datetime):
                if 0 < (t2 - t1).total_seconds() < 5:
                    sep = ", " if records[i]["Flags"] else ""
                    records[i]["Flags"]       += sep + "burst_activity"
                    records[i]["FlagReasons"] += (" | " if records[i]["FlagReasons"] else "") + \
                        "Consecutive events < 5 seconds apart — possible automated activity"
        except Exception:
            pass

    return records


# ══════════════════════════════════════════════════════════════════
#  CSV EXPORT
# ══════════════════════════════════════════════════════════════════

def export_csv(records: list[dict], path: str):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction="ignore")
        w.writeheader()
        w.writerows(records)
    print(f"[+] CSV exported  → {path}")


# ══════════════════════════════════════════════════════════════════
#  HTML INTERACTIVE REPORT  (v4 — with event detail panel)
# ══════════════════════════════════════════════════════════════════

def export_html(records: list[dict], path: str, case_name: str = "Case"):
    esc = html_lib.escape

    # Hourly histogram
    hourly = {}
    for rec in records:
        h = rec["Timestamp"][:13]
        hourly[h] = hourly.get(h, 0) + 1
    hs = sorted(hourly)

    flagged  = sum(1 for r in records if r.get("Flags"))
    total    = len(records)
    procs    = len(set(os.path.basename(r.get("ProcessPath","")) for r in records if r.get("ProcessPath")))
    urls     = len(set(r.get("URL","") for r in records if r.get("URL")))
    first_ts = records[0]["Timestamp"][:10]  if records else "N/A"
    last_ts  = records[-1]["Timestamp"][:10] if records else "N/A"
    now_str  = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Build JSON data blob for the detail panel (all fields per event)
    events_json = []
    for rec in records:
        events_json.append({
            "timestamp":   rec.get("Timestamp", ""),
            "windowTitle": rec.get("WindowTitle", ""),
            "processPath": rec.get("ProcessPath", ""),
            "url":         rec.get("URL", ""),
            "isForeground":rec.get("IsForeground", ""),
            "snapshotId":  rec.get("SnapshotId", ""),
            "appId":       rec.get("AppId", ""),
            "ocrText":     rec.get("OcrText", ""),
            "filename":    rec.get("Filename", ""),
            "fileSize":    rec.get("FileSize", 0),
            "sha256":      rec.get("SHA256", ""),
            "mtime":       rec.get("MTime", ""),
            "ctime":       rec.get("CTime", ""),
            "flags":       rec.get("Flags", ""),
            "flagReasons": rec.get("FlagReasons", ""),
            "source":      rec.get("Source", ""),
            "autopsy":     rec.get("AutopsyCorroboration", ""),
        })

    # Build table rows
    rows_html = ""
    for i, rec in enumerate(records):
        fc       = "flag-row" if rec.get("Flags") else ""
        proc_b   = esc(os.path.basename(rec.get("ProcessPath",""))[:38])
        sha_s    = (rec.get("SHA256","")[:14] + "…") if rec.get("SHA256") else "—"
        url_val  = rec.get("URL", "")
        url_disp = esc(url_val[:50]) if url_val else '<span class="no-url">—</span>'
        url_href = esc(url_val)
        title    = esc(str(rec.get("WindowTitle",""))[:72])
        fg_sym   = "✓" if str(rec.get("IsForeground","")).lower() in ("true","1","yes") else "○"
        flags    = esc(str(rec.get("Flags","")))
        src      = esc(rec.get("Source",""))

        rows_html += f"""
<tr class="{fc}" data-idx="{i}" onclick="showDetail({i})" style="cursor:pointer">
  <td>{esc(rec['Timestamp'])}</td>
  <td>{title}</td>
  <td><span class="proc">{proc_b}</span></td>
  <td>{"<a href='" + url_href + "' target='_blank' onclick='event.stopPropagation()'>" + url_disp + "</a>" if url_val else url_disp}</td>
  <td class="fg">{fg_sym}</td>
  <td class="flag-cell">{flags}</td>
  <td class="mono">{sha_s}</td>
  <td class="src">{src}</td>
</tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{esc(TOOL_NAME)} {esc(case_name)}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root{{
  --bg:#0d1117;--surf:#161b22;--surf2:#21262d;--surf3:#2d333b;
  --acc:#388bfd;--flag:#f85149;--ok:#3fb950;--warn:#e3b341;
  --txt:#c9d1d9;--txt2:#8b949e;--txt3:#6e7681;--bdr:#30363d;
  --font:'Segoe UI',system-ui,Arial,sans-serif;
}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:var(--bg);color:var(--txt);font-family:var(--font);font-size:13px}}
header{{background:var(--surf);border-bottom:2px solid var(--acc);padding:16px 28px}}
h1{{font-size:1.25rem;color:var(--acc);font-weight:700}}
.meta{{color:var(--txt2);font-size:.8rem;margin-top:4px}}
.stats{{display:flex;gap:10px;padding:14px 28px;flex-wrap:wrap}}
.card{{background:var(--surf);border:1px solid var(--bdr);border-radius:8px;
       padding:11px 16px;min-width:138px}}
.card .n{{font-size:1.8rem;font-weight:800;color:var(--acc)}}
.card .l{{color:var(--txt2);font-size:.75rem;margin-top:2px}}
.card.danger .n{{color:var(--flag)}}
.card.ok .n{{color:var(--ok)}}
.row2{{display:flex;gap:14px;margin:0 28px 18px;flex-wrap:wrap}}
.chart-box{{flex:2;min-width:300px;background:var(--surf);border-radius:8px;
            padding:14px;border:1px solid var(--bdr)}}
.chart-box h3,.pie-box h3{{color:var(--txt2);font-size:.78rem;
  text-transform:uppercase;letter-spacing:.06em;margin-bottom:9px}}
.pie-box{{flex:1;min-width:180px;background:var(--surf);border-radius:8px;
          padding:14px;border:1px solid var(--bdr)}}
.legend{{display:flex;flex-wrap:wrap;gap:8px;margin-top:9px;
         font-size:.73rem;color:var(--txt2)}}
.leg-sq{{width:9px;height:9px;border-radius:2px;display:inline-block;
         margin-right:3px;vertical-align:middle}}
.ctrl{{padding:0 28px 10px;display:flex;gap:10px;align-items:center;flex-wrap:wrap}}
.ctrl input{{background:var(--surf2);border:1px solid var(--bdr);color:var(--txt);
             padding:7px 12px;border-radius:6px;width:320px;font-size:.87rem}}
.ctrl input:focus{{outline:2px solid var(--acc)}}
.badge{{background:var(--acc);color:#fff;border-radius:10px;
        padding:2px 9px;font-size:.71rem}}
.btn-filter{{background:var(--surf2);border:1px solid var(--bdr);
             color:var(--txt2);padding:5px 11px;border-radius:6px;
             font-size:.78rem;cursor:pointer}}
.btn-filter.active{{background:var(--flag);color:#fff;border-color:var(--flag)}}
.tbl-wrap{{margin:0 28px 20px;overflow-x:auto;border-radius:8px;
           border:1px solid var(--bdr)}}
table{{width:100%;border-collapse:collapse}}
th{{background:var(--surf2);color:var(--txt2);text-transform:uppercase;
    font-size:.67rem;letter-spacing:.05em;padding:9px 11px;
    text-align:left;cursor:pointer;white-space:nowrap;user-select:none}}
th:hover{{color:var(--acc)}}
td{{padding:8px 11px;border-bottom:1px solid var(--bdr);
    vertical-align:middle;color:var(--txt)}}
tr:hover td{{background:var(--surf2)}}
tr.flag-row td{{background:rgba(248,81,73,.05)}}
tr.flag-row:hover td{{background:rgba(248,81,73,.11)}}
tr.selected td{{background:rgba(56,139,253,.12)!important;
                outline:1px solid var(--acc)}}
.flag-cell{{color:var(--flag);font-size:.75rem;font-weight:600}}
.mono{{font-family:monospace;font-size:.71rem;color:var(--txt2)}}
.proc{{background:var(--surf2);color:var(--txt2);border-radius:4px;
       padding:1px 6px;font-size:.74rem;font-family:monospace}}
.fg{{text-align:center;color:var(--ok)}}
.src{{font-size:.69rem;color:var(--txt3)}}
.no-url{{color:var(--txt3);font-size:.75rem}}
a{{color:var(--acc);text-decoration:none}}
a:hover{{text-decoration:underline}}
.hidden{{display:none}}

/* ── Detail panel ── */
#detail-panel{{
  position:fixed;right:0;top:0;bottom:0;width:420px;
  background:var(--surf);border-left:1px solid var(--bdr);
  transform:translateX(100%);transition:transform .22s ease;
  overflow-y:auto;z-index:100;display:flex;flex-direction:column;
}}
#detail-panel.open{{transform:translateX(0)}}
#detail-header{{
  background:var(--surf2);padding:14px 18px;
  border-bottom:1px solid var(--bdr);
  display:flex;align-items:center;justify-content:space-between;
  position:sticky;top:0;z-index:2;
}}
#detail-header h2{{font-size:.92rem;color:var(--acc);font-weight:600}}
#detail-close{{background:none;border:none;color:var(--txt2);
               font-size:1.2rem;cursor:pointer;padding:2px 6px;
               border-radius:4px}}
#detail-close:hover{{background:var(--surf3);color:var(--txt)}}
#detail-body{{padding:16px 18px;flex:1}}
.det-section{{margin-bottom:18px}}
.det-section h3{{font-size:.72rem;color:var(--txt2);text-transform:uppercase;
                 letter-spacing:.07em;margin-bottom:8px;
                 padding-bottom:5px;border-bottom:1px solid var(--bdr)}}
.det-row{{display:flex;gap:8px;margin-bottom:7px;align-items:flex-start}}
.det-label{{min-width:100px;color:var(--txt2);font-size:.75rem;flex-shrink:0;
            padding-top:1px}}
.det-val{{color:var(--txt);font-size:.8rem;word-break:break-all;flex:1}}
.det-val.mono{{font-family:monospace;font-size:.72rem}}
.det-val a{{color:var(--acc)}}
.flag-block{{background:rgba(248,81,73,.08);border:1px solid rgba(248,81,73,.25);
             border-radius:6px;padding:10px 12px;margin-bottom:8px}}
.flag-name{{color:var(--flag);font-weight:600;font-size:.8rem;
            font-family:monospace;margin-bottom:4px}}
.flag-reason{{color:var(--txt2);font-size:.77rem;line-height:1.45}}
.flag-reason em{{color:var(--warn);font-style:normal}}
.no-flags{{color:var(--ok);font-size:.8rem;}}
.corr-block{{background:rgba(63,185,80,.08);border:1px solid rgba(63,185,80,.25);
             border-radius:6px;padding:8px 12px;color:var(--ok);font-size:.78rem}}
.url-block{{background:var(--surf2);border-radius:6px;padding:8px 12px;
            font-family:monospace;font-size:.75rem;color:var(--acc);
            word-break:break-all}}
.url-none{{color:var(--txt3);font-size:.78rem}}
.url-why{{margin-top:6px;color:var(--txt2);font-size:.73rem;line-height:1.4}}
.hash-full{{font-family:monospace;font-size:.72rem;color:var(--txt2);
            word-break:break-all;background:var(--surf2);border-radius:4px;
            padding:6px 8px;margin-top:4px}}

footer{{text-align:center;padding:18px;color:var(--txt2);
        font-size:.75rem;border-top:1px solid var(--bdr)}}
</style>
</head>
<body>

<!-- ── Detail Panel ── -->
<div id="detail-panel">
  <div id="detail-header">
    <h2 id="dp-title">Event Details</h2>
    <button id="detail-close" onclick="closeDetail()" title="Close">✕</button>
  </div>
  <div id="detail-body"></div>
</div>

<header>
  <div>
    <h1>🔍 {esc(TOOL_NAME)} Forensic Report</h1>
    <div class="meta">
      Case: <strong style="color:var(--txt)">{esc(case_name)}</strong> &nbsp;|&nbsp;
      Generated: {now_str} &nbsp;|&nbsp;
      Authors: {esc(TOOL_AUTHOR)}
    </div>
  </div>
</header>

<div class="stats">
  <div class="card"><div class="n">{total}</div><div class="l">Total Events</div></div>
  <div class="card danger"><div class="n">{flagged}</div><div class="l">Flagged Events</div></div>
  <div class="card ok"><div class="n">{total-flagged}</div><div class="l">Clean Events</div></div>
  <div class="card"><div class="n">{procs}</div><div class="l">Unique Processes</div></div>
  <div class="card"><div class="n">{urls}</div><div class="l">Unique URLs</div></div>
  <div class="card"><div class="n">{first_ts}</div><div class="l">First Event</div></div>
  <div class="card"><div class="n">{last_ts}</div><div class="l">Last Event</div></div>
</div>

<div class="row2">
  <div class="chart-box">
    <h3>Activity Frequency — Events per Hour</h3>
    <canvas id="actChart" height="88"></canvas>
  </div>
  <div class="pie-box">
    <h3>Process Distribution</h3>
    <canvas id="pieChart" height="145"></canvas>
    <div class="legend" id="pieLegend"></div>
  </div>
</div>

<div class="ctrl">
  <input id="sb" type="text"
    placeholder="🔎  Search timestamp, window, process, URL, flags…"
    oninput="filt()">
  <button class="btn-filter" id="btnFlagged" onclick="toggleFlagged()">
    ⚑ Show flagged only
  </button>
  <span class="badge" id="cnt">{total} events</span>
  <span style="color:var(--txt2);font-size:.75rem;margin-left:4px">
    ↓ Click any row for details
  </span>
</div>

<div class="tbl-wrap">
<table>
<thead>
<tr>
  <th onclick="srt(0)">Timestamp ↕</th>
  <th onclick="srt(1)">Window Title ↕</th>
  <th onclick="srt(2)">Process ↕</th>
  <th>URL</th>
  <th>FG</th>
  <th onclick="srt(5)">Flags ↕</th>
  <th>SHA-256</th>
  <th>Source</th>
</tr>
</thead>
<tbody id="tb">{rows_html}</tbody>
</table>
</div>

<footer>
  {esc(TOOL_NAME)} For authorized forensic use only.
  Chain of custody must be maintained at all times.
</footer>

<script>
const EVENTS = {json.dumps(events_json, ensure_ascii=False)};
const COLORS = ['#388bfd','#3fb950','#f85149','#e3b341','#a371f7','#f78166','#58a6ff','#56d364'];

// ── Charts ────────────────────────────────────────────────────────
const ctx1 = document.getElementById('actChart').getContext('2d');
new Chart(ctx1, {{type:'bar',data:{{
  labels:{json.dumps(hs)},
  datasets:[{{label:'Events',data:{json.dumps([hourly[h] for h in hs])},
    backgroundColor:'rgba(56,139,253,.5)',borderColor:'rgba(56,139,253,1)',borderWidth:1}}]
}},options:{{responsive:true,plugins:{{legend:{{display:false}}}},
  scales:{{
    x:{{ticks:{{color:'#8b949e',font:{{size:9}},maxRotation:30}},grid:{{color:'#21262d'}}}},
    y:{{ticks:{{color:'#8b949e'}},grid:{{color:'#21262d'}},beginAtZero:true}}
  }}
}}}});

const procMap={{}};
document.querySelectorAll('#tb tr').forEach(r=>{{
  const p=r.cells[2].textContent.trim();
  if(p) procMap[p]=(procMap[p]||0)+1;
}});
const pL=Object.keys(procMap), pD=pL.map(k=>procMap[k]);
new Chart(document.getElementById('pieChart').getContext('2d'),{{
  type:'doughnut',
  data:{{labels:pL,datasets:[{{data:pD,
    backgroundColor:COLORS.slice(0,pL.length),borderWidth:2,borderColor:'#161b22'}}]}},
  options:{{responsive:true,plugins:{{legend:{{display:false}}}},cutout:'58%'}}
}});
const leg=document.getElementById('pieLegend');
pL.forEach((l,i)=>{{
  leg.innerHTML+=`<span><span class="leg-sq" style="background:${{COLORS[i%COLORS.length]}}"></span>${{l}} (${{pD[i]}})</span>`;
}});

// ── Search & filter ───────────────────────────────────────────────
let showFlaggedOnly=false;
function filt(){{
  const q=document.getElementById('sb').value.toLowerCase();
  let v=0;
  document.querySelectorAll('#tb tr').forEach(r=>{{
    const flagged=r.classList.contains('flag-row');
    const textMatch=!q||r.textContent.toLowerCase().includes(q);
    const show=textMatch&&(!showFlaggedOnly||flagged);
    r.classList.toggle('hidden',!show);
    if(show)v++;
  }});
  document.getElementById('cnt').textContent=v+' events';
}}
function toggleFlagged(){{
  showFlaggedOnly=!showFlaggedOnly;
  document.getElementById('btnFlagged').classList.toggle('active',showFlaggedOnly);
  filt();
}}

// ── Sort ──────────────────────────────────────────────────────────
function srt(c){{
  const tb=document.getElementById('tb');
  const rows=[...tb.querySelectorAll('tr')];
  const asc=tb.dataset.c==c&&tb.dataset.d=='a';
  rows.sort((a,b)=>{{
    const av=a.cells[c]?.textContent.trim()||'';
    const bv=b.cells[c]?.textContent.trim()||'';
    return asc?bv.localeCompare(av):av.localeCompare(bv);
  }});
  tb.dataset.c=c; tb.dataset.d=asc?'d':'a';
  rows.forEach(r=>tb.appendChild(r));
}}

// ── Detail panel ─────────────────────────────────────────────────
let currentIdx=null;
function showDetail(idx){{
  const e=EVENTS[idx];
  if(!e) return;

  // De-select previous, select new
  document.querySelectorAll('#tb tr.selected').forEach(r=>r.classList.remove('selected'));
  const row=document.querySelector(`#tb tr[data-idx="${{idx}}"]`);
  if(row) row.classList.add('selected');
  currentIdx=idx;

  // Build flag section
  let flagHtml='';
  if(e.flags){{
    const flagList=e.flags.split(',').map(f=>f.trim()).filter(Boolean);
    const reasonList=e.flagReasons.split('|').map(r=>r.trim()).filter(Boolean);
    flagList.forEach((f,i)=>{{
      const reason=reasonList[i]||'Suspicious keyword detected';
      flagHtml+=`<div class="flag-block">
        <div class="flag-name">⚑ ${{f.toUpperCase()}}</div>
        <div class="flag-reason">${{reason}}</div>
      </div>`;
    }});
  }} else {{
    flagHtml=`<div class="no-flags">✓ No suspicious indicators detected</div>`;
  }}

  // URL section
  let urlHtml='';
  if(e.url){{
    urlHtml=`<div class="url-block"><a href="${{e.url}}" target="_blank">${{e.url}}</a></div>`;
  }} else {{
    urlHtml=`<div class="url-none">No URL recorded for this event.</div>
    <div class="url-why">
      URLs are only captured when the active window is a web browser and
      Windows Recall can read the address bar. For non-browser applications
      (File Explorer, Terminal, Office) no URL is stored — this is expected
      behaviour, not a data loss issue.
    </div>`;
  }}

  // Autopsy section
  const autopsyHtml = e.autopsy
    ? `<div class="corr-block">✓ ${{e.autopsy}}</div>`
    : `<span style="color:var(--txt3);font-size:.78rem">No filesystem corroboration found</span>`;

  document.getElementById('dp-title').textContent=
    e.windowTitle.slice(0,45)||(e.filename||'Event Details');

  document.getElementById('detail-body').innerHTML=`
    <div class="det-section">
      <h3>Event Identity</h3>
      <div class="det-row"><span class="det-label">Timestamp</span>
        <span class="det-val mono">${{e.timestamp}}</span></div>
      <div class="det-row"><span class="det-label">Source</span>
        <span class="det-val">${{e.source}}</span></div>
      <div class="det-row"><span class="det-label">Snapshot ID</span>
        <span class="det-val mono">${{e.snapshotId||'—'}}</span></div>
      <div class="det-row"><span class="det-label">App ID</span>
        <span class="det-val">${{e.appId||'—'}}</span></div>
    </div>

    <div class="det-section">
      <h3>Activity</h3>
      <div class="det-row"><span class="det-label">Window Title</span>
        <span class="det-val">${{e.windowTitle||'—'}}</span></div>
      <div class="det-row"><span class="det-label">Process Path</span>
        <span class="det-val mono">${{e.processPath||'—'}}</span></div>
      <div class="det-row"><span class="det-label">Foreground</span>
        <span class="det-val">${{e.isForeground==='true'||e.isForeground==='1'?'✓ Yes':'✗ No'}}</span></div>
    </div>

    <div class="det-section">
      <h3>URL</h3>
      ${{urlHtml}}
    </div>

    ${{e.ocrText?`<div class="det-section"><h3>OCR Text</h3>
      <div style="font-size:.78rem;color:var(--txt2);line-height:1.5;
        background:var(--surf2);padding:8px 10px;border-radius:6px;">
        ${{e.ocrText}}</div></div>`:''}}`+`

    <div class="det-section">
      <h3>Flags &amp; Reasons</h3>
      ${{flagHtml}}
    </div>

    <div class="det-section">
      <h3>File Integrity</h3>
      <div class="det-row"><span class="det-label">Filename</span>
        <span class="det-val mono">${{e.filename||'—'}}</span></div>
      <div class="det-row"><span class="det-label">File Size</span>
        <span class="det-val">${{e.fileSize? (e.fileSize/1024).toFixed(1)+' KB' : '—'}}</span></div>
      <div class="det-row"><span class="det-label">Modified</span>
        <span class="det-val mono">${{e.mtime||'—'}}</span></div>
      <div class="det-row"><span class="det-label">Created</span>
        <span class="det-val mono">${{e.ctime||'—'}}</span></div>
      <div style="margin-top:6px;font-size:.72rem;color:var(--txt2)">SHA-256</div>
      <div class="hash-full">${{e.sha256||'—'}}</div>
    </div>

    <div class="det-section">
      <h3>Autopsy Corroboration</h3>
      ${{autopsyHtml}}
    </div>
  `;

  document.getElementById('detail-panel').classList.add('open');
}}

function closeDetail(){{
  document.getElementById('detail-panel').classList.remove('open');
  document.querySelectorAll('#tb tr.selected').forEach(r=>r.classList.remove('selected'));
  currentIdx=null;
}}

document.addEventListener('keydown',e=>{{
  if(e.key==='Escape') closeDetail();
}});
</script>
</body>
</html>"""

    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[+] HTML exported → {path}")


# ══════════════════════════════════════════════════════════════════
#  PDF EXPORT (Professional, colored, using fpdf2)
# ══════════════════════════════════════════════════════════════════

def export_pdf(records: list[dict], path: str, case_name: str):
    if not PDF_AVAILABLE:
        print("  [WARN] fpdf2 not installed, skipping PDF export. Install with: pip install fpdf2")
        return

    # ── Color palette ──
    PRIMARY     = (26, 54, 93)      # Deep navy
    ACCENT      = (41, 128, 185)    # Steel blue
    DANGER      = (192, 57, 43)     # Crimson red
    DANGER_LITE = (253, 237, 236)   # Light red tint
    SUCCESS     = (39, 174, 96)     # Forest green
    SUCCESS_LITE= (232, 248, 239)   # Light green tint
    DARK        = (30, 30, 30)      # Near-black text
    MID_GRAY    = (108, 117, 125)   # Muted gray
    LIGHT_GRAY  = (222, 226, 230)   # Table alt row
    SILVER      = (245, 246, 248)   # Page background tint
    WHITE       = (255, 255, 255)

    # ── Helper to sanitize text ──
    def safe_text(text):
        if not text:
            return ""
        replacements = {
            '\u2014': '--', '\u2013': '-', '\u2018': "'",
            '\u2019': "'", '\u201c': '"', '\u201d': '"',
            '\u2026': '...', '\u2605': '*', '\u2713': 'Y',
        }
        for char, replacement in replacements.items():
            text = text.replace(char, replacement)
        return text.encode('latin-1', errors='replace').decode('latin-1')

    generated_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # ── FPDF subclass for persistent header & footer on every page ──
    class ForensicPDF(FPDF):
        def header(self):
            # Top colour bar (full width, no margins)
            self.set_fill_color(*PRIMARY)
            self.rect(0, 0, self.w, 18, 'F')

            # Thin accent stripe
            self.set_fill_color(*ACCENT)
            self.rect(0, 18, self.w, 2, 'F')

            # Tool name — left-aligned inside the bar
            self.set_xy(12, 4)
            self.set_font("Helvetica", "B", 11)
            self.set_text_color(*WHITE)
            self.cell(80, 6, safe_text(f"{TOOL_NAME} Forensic Report"), border=0)

            # Case name — centre
            self.set_xy(0, 4)
            self.set_font("Helvetica", "", 9)
            self.set_text_color(180, 210, 240)
            self.cell(self.w, 6, safe_text(f"Case: {case_name}"), border=0, align="C")

            # Page number — right
            self.set_xy(self.w - 55, 4)
            self.set_font("Helvetica", "", 8)
            self.set_text_color(160, 195, 230)
            self.cell(43, 6, f"Page {self.page_no()}", border=0, align="R")

            # Reset cursor below header
            self.set_text_color(*DARK)
            self.set_y(26)

        def footer(self):
            # Footer separator line
            self.set_y(-16)
            self.set_draw_color(*ACCENT)
            self.set_line_width(0.4)
            self.line(12, self.get_y(), self.w - 12, self.get_y())
            self.ln(1)

            # Left: authors
            self.set_font("Helvetica", "", 7)
            self.set_text_color(*MID_GRAY)
            self.set_x(12)
            self.cell(90, 5, safe_text(f"Authors: {TOOL_AUTHOR}"), border=0)

            # Centre: classification
            self.set_x(0)
            self.set_text_color(*DANGER)
            self.set_font("Helvetica", "B", 7)
            self.cell(self.w, 5, "CONFIDENTIAL -- For Authorized Forensic Use Only",
                      border=0, align="C")

            # Right: generated timestamp
            self.set_x(self.w - 62)
            self.set_font("Helvetica", "", 7)
            self.set_text_color(*MID_GRAY)
            self.cell(50, 5, safe_text(f"Generated: {generated_at}"), border=0, align="R")

    pdf = ForensicPDF()
    pdf.set_margins(12, 26, 12)
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # ── Cover / title block (first page only) ──
    # Dark banner panel
    pdf.set_fill_color(*PRIMARY)
    pdf.rect(12, 28, pdf.w - 24, 38, 'F')

    pdf.set_xy(12, 34)
    pdf.set_font("Helvetica", "B", 22)
    pdf.set_text_color(*WHITE)
    pdf.cell(pdf.w - 24, 12, safe_text(f"{TOOL_NAME}"), border=0, align="C",
             new_x="LMARGIN", new_y="NEXT")

    pdf.set_x(12)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(170, 200, 235)
    pdf.cell(pdf.w - 24, 7, "Windows Recall Forensic Analysis Report",
             border=0, align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.set_x(12)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(140, 180, 220)
    pdf.cell(pdf.w - 24, 5,
             safe_text(f"Case: {case_name}   |   {generated_at}"),
             border=0, align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.set_text_color(*DARK)
    pdf.ln(12)

    # ── Section heading helper ──
    def section_heading(label, color):
        if pdf.get_y() > 240:
            pdf.add_page()
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(*color)
        # Left accent bar
        bar_x = pdf.l_margin
        bar_y = pdf.get_y()
        pdf.set_fill_color(*color)
        pdf.rect(bar_x, bar_y, 3, 7, 'F')
        pdf.set_x(bar_x + 5)
        pdf.cell(0, 7, safe_text(label), new_x="LMARGIN", new_y="NEXT")
        # Underline
        pdf.set_draw_color(*color)
        pdf.set_line_width(0.3)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
        pdf.set_line_width(0.2)
        pdf.ln(4)
        pdf.set_text_color(*DARK)

    # ── Summary cards ──
    total    = len(records)
    flagged  = sum(1 for r in records if r.get("Flags"))
    procs    = len(set(os.path.basename(r.get("ProcessPath","")) for r in records if r.get("ProcessPath")))
    urls     = len(set(r.get("URL","") for r in records if r.get("URL")))
    corroborated = sum(1 for r in records if r.get("AutopsyCorroboration"))
    first_ts = records[0]["Timestamp"][:10] if records else "N/A"
    last_ts  = records[-1]["Timestamp"][:10] if records else "N/A"

    section_heading("Case Summary", PRIMARY)

    card_w   = 56
    card_h   = 18
    col_gap  = 5
    start_x  = pdf.l_margin
    start_y  = pdf.get_y()

    card_data = [
        ("Total Events",    str(total),              PRIMARY,  SILVER),
        ("Flagged",         str(flagged),             DANGER,   DANGER_LITE),
        ("Clean",           str(total - flagged),     SUCCESS,  SUCCESS_LITE),
        ("Unique Processes",str(procs),               ACCENT,   SILVER),
        ("Unique URLs",     str(urls),                ACCENT,   SILVER),
        ("Corroborated",    f"{corroborated}/{total}",SUCCESS,  SUCCESS_LITE),
        ("Time Span",       f"{first_ts} to {last_ts}", PRIMARY, SILVER),
    ]

    for i, (label, value, txt_color, bg_color) in enumerate(card_data):
        col = i % 3
        row = i // 3
        x   = start_x + col * (card_w + col_gap)
        y   = start_y + row * (card_h + 5)

        # Card background
        pdf.set_xy(x, y)
        pdf.set_fill_color(*bg_color)
        pdf.set_draw_color(*LIGHT_GRAY)
        pdf.rect(x, y, card_w, card_h, 'FD')

        # Left accent stripe on card
        pdf.set_fill_color(*txt_color)
        pdf.rect(x, y, 3, card_h, 'F')

        # Value (large)
        pdf.set_xy(x + 5, y + 2)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(*txt_color)
        pdf.cell(card_w - 6, 8, safe_text(value), border=0, align="L")

        # Label (small, muted)
        pdf.set_xy(x + 5, y + 10)
        pdf.set_font("Helvetica", "", 7)
        pdf.set_text_color(*MID_GRAY)
        pdf.cell(card_w - 6, 5, safe_text(label), border=0, align="L")

    rows_used = -(-len(card_data) // 3)   # ceiling div
    pdf.set_y(start_y + rows_used * (card_h + 5) + 6)
    pdf.set_text_color(*DARK)

    # ── Flagged Events Table ──
    flagged_records = [r for r in records if r.get("Flags")]
    if flagged_records:
        section_heading("Flagged Events", DANGER)

        effective_w = pdf.w - pdf.l_margin - pdf.r_margin
        col_widths  = [44, 80, 56, 10]
        headers     = ["Timestamp", "Window Title", "Flags", "FG"]

        # Table header row
        pdf.set_fill_color(*PRIMARY)
        pdf.set_text_color(*WHITE)
        pdf.set_font("Helvetica", "B", 8)
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 7, safe_text(h), border=0, fill=True, align="C")
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        alt_fill = False
        for rec in flagged_records:
            if pdf.get_y() > 262:
                pdf.add_page()
                # Repeat header
                pdf.set_fill_color(*PRIMARY)
                pdf.set_text_color(*WHITE)
                pdf.set_font("Helvetica", "B", 8)
                for i, h in enumerate(headers):
                    pdf.cell(col_widths[i], 7, safe_text(h), border=0, fill=True, align="C")
                pdf.ln()
                pdf.set_font("Helvetica", "", 8)

            ts    = safe_text(rec["Timestamp"][:19])
            title = safe_text(rec["WindowTitle"][:64])
            flags = safe_text(rec["Flags"][:48])
            fg    = "Y" if str(rec.get("IsForeground","")).lower() in ("true","1","yes") else "N"

            bg = DANGER_LITE if alt_fill else WHITE
            pdf.set_fill_color(*bg)
            pdf.set_text_color(*DARK)
            pdf.cell(col_widths[0], 5.5, ts,    border=0, fill=True)
            pdf.cell(col_widths[1], 5.5, title, border=0, fill=True)
            # Flag cell — colour code FG
            fg_color = DANGER if fg == "Y" else MID_GRAY
            pdf.set_text_color(*fg_color if flags else MID_GRAY)
            pdf.cell(col_widths[2], 5.5, flags, border=0, fill=True)
            pdf.set_text_color(*DANGER if fg == "Y" else DARK)
            pdf.set_font("Helvetica", "B" if fg == "Y" else "", 8)
            pdf.cell(col_widths[3], 5.5, fg,    border=0, fill=True, align="C")
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(*DARK)
            pdf.ln()

            alt_fill = not alt_fill

        # Bottom border line
        pdf.set_draw_color(*LIGHT_GRAY)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
        pdf.ln(7)

    # ── Autopsy Corroboration ──
    corroborated_records = [r for r in records if r.get("AutopsyCorroboration")]
    if corroborated_records:
        section_heading("Autopsy Corroboration", SUCCESS)

        effective_w = pdf.w - pdf.l_margin - pdf.r_margin
        for rec in corroborated_records:
            if pdf.get_y() > 262:
                pdf.add_page()
            ts    = safe_text(rec["Timestamp"][:19])
            title = safe_text(rec["WindowTitle"][:55])
            corr  = safe_text(rec.get("AutopsyCorroboration", ""))

            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(*DARK)
            pdf.cell(0, 5, f"{ts}  |  {title}", new_x="LMARGIN", new_y="NEXT")

            pdf.set_fill_color(*SUCCESS_LITE)
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(*SUCCESS)
            pdf.multi_cell(w=effective_w, h=4, text=f"    {corr}", fill=True)
            pdf.set_text_color(*DARK)
            pdf.ln(2)
        pdf.ln(5)

    # ── File Integrity ──
    effective_w = pdf.w - pdf.l_margin - pdf.r_margin
    section_heading("File Integrity (SHA-256)", PRIMARY)

    alt_fill = False
    for rec in records:
        sha = rec.get("SHA256", "")
        if sha:
            if pdf.get_y() > 262:
                pdf.add_page()

            fname  = safe_text(rec.get("Filename", "unknown"))
            bg     = SILVER if alt_fill else WHITE
            alt_fill = not alt_fill

            pdf.set_fill_color(*bg)
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(*ACCENT)
            pdf.cell(effective_w, 4.5, fname, new_x="LMARGIN", new_y="NEXT", fill=True)

            pdf.set_font("Courier", "", 7)
            pdf.set_text_color(80, 80, 80)
            pdf.multi_cell(w=effective_w, h=3.5, text=f"  {sha}", fill=True)
            pdf.ln(1)

    pdf.ln(5)

    # ── Chain of Custody ──
    if pdf.get_y() > 220:
        pdf.add_page()
    section_heading("Chain of Custody", PRIMARY)

    pdf.set_fill_color(*SILVER)
    pdf.set_draw_color(*LIGHT_GRAY)
    coc_text = (
        "This report was generated using RecallTimeline, an open-source Windows Recall forensic tool. "
        "All evidence items referenced herein are preserved with cryptographic hashes (SHA-256) to ensure integrity. "
        "Any alteration or tampering will invalidate these hashes."
    )
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*DARK)
    pdf.multi_cell(w=effective_w, h=5, text=safe_text(coc_text), fill=True, border=1)
    pdf.ln(8)

    # Signature block
    sig_col = effective_w / 2 - 4
    for label in ("Examiner", "Reviewer"):
        x = pdf.get_x()
        y = pdf.get_y()
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*MID_GRAY)
        pdf.cell(sig_col, 5, safe_text(label + ":"), border=0)
        pdf.cell(8, 5, "", border=0)
        pdf.ln(5)
        # Signature line
        pdf.set_draw_color(*LIGHT_GRAY)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.l_margin + sig_col - 10, pdf.get_y())
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(*MID_GRAY)
        pdf.set_x(pdf.l_margin)
        pdf.cell(sig_col - 10, 4, "Signature", border=0)
        pdf.cell(20, 4, "Date: ____________", border=0)
        pdf.ln(10)

    try:
        pdf.output(path)
        print(f"[+] PDF exported  -> {path}")
    except Exception as e:
        print(f"  [WARN] PDF generation failed: {e}")


# ══════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════

def cmd_analyse(args):
    print(BANNER)
    os.makedirs(args.output_dir, exist_ok=True)

    records = build_timeline(args.recall_dir, args.autopsy_csv)
    if not records:
        print("[!] No parseable artifacts found. Exiting.")
        sys.exit(1)

    if not args.no_flag:
        records = flag_anomalies(records)

    base = os.path.join(args.output_dir, args.case.replace(" ", "_"))
    export_csv (records, f"{base}_timeline.csv")
    export_html(records, f"{base}_timeline.html", case_name=args.case)
    export_pdf (records, f"{base}_timeline.pdf",  case_name=args.case)

    flagged = sum(1 for r in records if r.get("Flags"))
    print(f"\n{'='*60}")
    print(f"  [✓] Complete — {len(records)} events, {flagged} flagged")
    print(f"  Output: {args.output_dir}")
    print(f"  Files : {args.case}_timeline.csv")
    print(f"          {args.case}_timeline.html")
    if PDF_AVAILABLE:
        print(f"          {args.case}_timeline.pdf")
    print(f"{'='*60}\n")


def main():
    parser = argparse.ArgumentParser(
        prog=TOOL_NAME,
        description=f"{TOOL_NAME} Windows Recall Forensic Timeline Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Analyse real Windows Recall folder
  python RecallTimeline.py analyse \\
      --recall-dir "C:\\Users\\USER\\AppData\\Local\\CoreAIPlatform.00\\UKP\\{GUID}"

  # Analyse capture_snapshots.py output
  python RecallTimeline.py analyse --recall-dir ./snapshots --case MyCapture
        """
    )
    
    sub = parser.add_subparsers(dest="command", required=True)
    
    p_an = sub.add_parser("analyse", help="Analyse a Recall artifact folder")
    p_an.add_argument("--recall-dir", required=True,
                      help="Path to folder with ukg.db, ImageStore/, or JPEG snapshots")
    p_an.add_argument("--output-dir",  default="./recall_output")
    p_an.add_argument("--case",        default="Investigation")
    p_an.add_argument("--autopsy-csv", default=None)
    p_an.add_argument("--no-flag",     action="store_true")
    p_an.set_defaults(func=cmd_analyse)
    
    args = parser.parse_args()
    args.func(args)

if __name__ == "__main__":
    main()
